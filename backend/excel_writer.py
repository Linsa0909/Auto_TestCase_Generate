import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ExcelWriter:
    """Generate Excel test case files matching the DevOps import template."""

    HEADERS = [
        "用例分组",
        "用例标题",
        "用例类型",
        "重要程度",
        "前置条件",
        "步骤描述",
        "预期结果",
        "需求编号(多个以逗号隔开)",
        "需求名称",
    ]

    COLUMN_WIDTHS = [16, 44, 14, 10, 48, 55, 55, 24, 20]

    @staticmethod
    def _normalize_requirement_id(rid: str) -> str:
        """Normalize requirement ID: strip # prefix, ensure trailing comma."""
        if not rid:
            return ""
        rid = rid.strip().lstrip('#')
        if rid and not rid.endswith(','):
            rid = rid + ','
        return rid

    def write(self, test_cases: list, filepath: str,
              requirement_name: str = "") -> str:
        """
        Write test cases to an Excel file.

        Args:
            test_cases: List of test case dicts with 'steps' sub-list
            filepath: Output file path
            requirement_name: Default requirement name if not set per case

        Returns:
            The filepath of the generated Excel file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet0"

        # Styles
        header_font = Font(name="Microsoft YaHei", bold=True, size=11)
        header_fill = PatternFill(
            start_color="F1F5F9", end_color="F1F5F9", fill_type="solid",
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # Write headers
        for col, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        row = 2
        for tc in test_cases:
            steps = tc.get("steps", [{}])
            if not steps:
                steps = [{}]

            for i, step in enumerate(steps):
                r = row + i
                if i == 0:
                    # First step row: fill all columns
                    ws.cell(row=r, column=1, value=tc.get("group", "")).alignment = cell_alignment
                    ws.cell(row=r, column=2, value=tc.get("title", "")).alignment = cell_alignment
                    ws.cell(row=r, column=3, value=tc.get("type", "手动测试用例")).alignment = cell_alignment
                    ws.cell(row=r, column=4, value=tc.get("priority", "L1")).alignment = cell_alignment
                    ws.cell(row=r, column=5, value=tc.get("precondition", "")).alignment = cell_alignment
                    rid = self._normalize_requirement_id(tc.get("requirement_id", ""))
                    ws.cell(row=r, column=8, value=rid).alignment = cell_alignment
                    ws.cell(row=r, column=9, value=tc.get("requirement_name", "") or requirement_name).alignment = cell_alignment
                    # Apply border to all cells in first row
                    for c in range(1, 10):
                        ws.cell(row=r, column=c).border = thin_border

                # Every step row: fill step and expected
                ws.cell(row=r, column=6, value=step.get("step", "")).alignment = cell_alignment
                ws.cell(row=r, column=7, value=step.get("expected", "")).alignment = cell_alignment
                if i > 0:
                    for c in range(1, 10):
                        ws.cell(row=r, column=c).border = thin_border

            row += len(steps)

        # Set column widths
        for i, width in enumerate(self.COLUMN_WIDTHS, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:I{row - 1}"

        # Ensure parent directory exists
        os.makedirs(os.path.dirname(filepath) if os.path.dirname(filepath) else ".", exist_ok=True)
        wb.save(filepath)
        return filepath

    @staticmethod
    def read(filepath: str) -> list:
        """
        Read test cases back from an Excel file.
        Returns a list of test case dicts compatible with the frontend DataTable.
        """
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        test_cases = []
        current_tc = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = (row[1] or "").strip() if len(row) > 1 else ""
            step_text = (row[5] or "").strip() if len(row) > 5 else ""
            expected_text = (row[6] or "").strip() if len(row) > 6 else ""

            if title:
                # New test case
                raw_type_map = {"L0": "冒烟", "L1": "功能", "L2": "边界"}
                priority = (row[3] or "L1").strip() if len(row) > 3 else "L1"
                current_tc = {
                    "group": (row[0] or "").strip() if len(row) > 0 else "",
                    "title": title,
                    "type": (row[2] or "手动测试用例").strip() if len(row) > 2 else "手动测试用例",
                    "priority": priority,
                    "precondition": (row[4] or "").strip() if len(row) > 4 else "",
                    "requirement_id": (row[7] or "").strip() if len(row) > 7 else "",
                    "requirement_name": (row[8] or "").strip() if len(row) > 8 else "",
                    "raw_type": raw_type_map.get(priority, "功能"),
                    "raw_level": "高" if priority == "L0" else "中" if priority == "L1" else "低",
                    "steps": [],
                }
                test_cases.append(current_tc)

            if current_tc is not None and (step_text or expected_text):
                current_tc["steps"].append({
                    "step": step_text,
                    "expected": expected_text,
                })

        wb.close()
        return test_cases
